#!/usr/bin/env python3
"""
AEGIS PulsePoint — Dashboard Data Updater
==========================================
Reads the weekly Project Portfolio .xlsx export (15-column simplified schema)
and writes dashboard_data.json for the single-page dashboard.

Usage:
    python update_dashboard.py <path_to_xlsx> [--output dashboard_data.json]

Expected columns:
    New | MTP | Program / Office | Subprog | Project or Task | Health |
    Status | Project Name | Coordinator | Start Date | End Date |
    AGDT Flag | Latest Update | Lookahead | XM - TR
"""

import argparse
import json
import sys
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required.  pip install openpyxl")


COLUMN_MAP = {
    "New":              "new",
    "MTP":              "mtp",
    "Program / Office": "program",
    "Subprog":          "subprogram",
    "Project or Task":  "project_or_task",
    "Health":           "health",
    "Status":           "status",
    "Project Name":     "project_name",
    "Coordinator":      "coordinator",
    "Start Date":       "start_date",
    "End Date":         "end_date",
    "AGDT Flag":        "agdt_flag",
    "Latest Update":    "latest_update",
    "Lookahead":        "lookahead",
    "XM - TR":          "xm_tr",
}


def serialize(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        return round(val, 6)
    if isinstance(val, str) and val.startswith("#"):
        return None
    return val


def read_portfolio(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Project Portfolio" in wb.sheetnames:
        ws = wb["Project Portfolio"]
    else:
        ws = wb[wb.sheetnames[0]]
        print(f"  ⚠  Sheet 'Project Portfolio' not found; using '{ws.title}'")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        sys.exit("Error: spreadsheet is empty.")

    header = rows[0]
    col_index = {}
    for i, cell in enumerate(header):
        if cell and str(cell).strip() in COLUMN_MAP:
            col_index[COLUMN_MAP[str(cell).strip()]] = i

    missing = set(COLUMN_MAP.values()) - set(col_index.keys())
    if missing:
        print(f"  ⚠  Missing columns (will be null): {', '.join(sorted(missing))}")

    projects = []
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        name_idx = col_index.get("project_name")
        if name_idx is not None and row[name_idx] is None:
            continue

        record = {}
        for json_key, col_i in col_index.items():
            record[json_key] = serialize(row[col_i]) if col_i < len(row) else None
        projects.append(record)

    return projects


def compute_summary(projects: list[dict]) -> dict:
    total = len(projects)
    health_counts = {"Green": 0, "Yellow": 0, "Red": 0, "Unknown": 0}
    status_counts = {}
    program_counts = {}
    coordinator_counts = {}

    for p in projects:
        h = (p.get("health") or "Unknown").strip().title()
        health_counts[h] = health_counts.get(h, 0) + 1

        s = (p.get("status") or "Unknown").strip()
        status_counts[s] = status_counts.get(s, 0) + 1

        prog = (p.get("program") or "Unassigned").strip()
        program_counts[prog] = program_counts.get(prog, 0) + 1

        coord = (p.get("coordinator") or "Unassigned").strip()
        coordinator_counts[coord] = coordinator_counts.get(coord, 0) + 1

    return {
        "total_projects": total,
        "health": health_counts,
        "statuses": status_counts,
        "programs": program_counts,
        "coordinators": coordinator_counts,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Convert Project Portfolio XLSX → JSON for PulsePoint dashboard."
    )
    parser.add_argument("xlsx", help="Path to the weekly .xlsx export")
    parser.add_argument("--output", "-o", default="dashboard_data.json",
                        help="Output JSON file path (default: dashboard_data.json)")
    parser.add_argument("--embed", "-e", default=None, metavar="HTML_FILE",
                        help="Embed data directly into the HTML file (no server needed)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"Error: file not found → {xlsx_path}")

    print(f"📂  Reading {xlsx_path.name} …")
    projects = read_portfolio(str(xlsx_path))
    print(f"   Found {len(projects)} project(s).")

    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": xlsx_path.name,
        "summary": compute_summary(projects),
        "projects": projects,
    }

    json_str = json.dumps(payload, indent=2, ensure_ascii=False)

    # Always write the standalone JSON
    out_path = Path(args.output)
    out_path.write_text(json_str)
    print(f"✅  Wrote {out_path}  ({out_path.stat().st_size:,} bytes)")

    # Optionally embed into HTML for static file:// usage
    if args.embed:
        html_path = Path(args.embed)
        if not html_path.exists():
            sys.exit(f"Error: HTML file not found → {html_path}")

        html = html_path.read_text(encoding="utf-8")

        # Build the script tag to inject
        data_script = (
            '<script>window.__PULSEPOINT_DATA__ = '
            + json.dumps(payload, ensure_ascii=False)
            + ';</script>'
        )

        # Remove any previously embedded data block
        import re
        html = re.sub(
            r'<script>window\.__PULSEPOINT_DATA__\s*=\s*.*?;</script>\s*',
            '',
            html,
            flags=re.DOTALL,
        )

        # Insert right before the closing </head> tag
        html = html.replace('</head>', data_script + '\n</head>', 1)

        html_path.write_text(html, encoding="utf-8")
        print(f"✅  Embedded data into {html_path}")
        print(f"   → Open {html_path} directly in your browser — no server needed.")


if __name__ == "__main__":
    main()
